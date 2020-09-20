

# %%
import numpy as np
import openpyxl as oxl
import math
import xlwt as xt
import pandas as pd


# 路径
tar_path = r'C:\Users\wuziyang\Desktop\数模题\附件一：325个样本数据.xlsx'
del_path = r'C:\Users\wuziyang\Desktop\corr_del_sample.xlsx'
ori_path = r'C:\Users\wuziyang\Desktop\数模题\附件三：285号和313号样本原始数据.xlsx'


# 获取目标数据
def load_tar_xml(path):
    # 操作变量列
    op_cols = []
    op_data = []
    # 性质列
    pro_cols = []
    pro_data = []

    wb = oxl.load_workbook(path)
    tar_sheet = wb.get_sheet_by_name("Sheet1")

    # 读取文件得到的行数不正确，手写
    row = 328
    col = tar_sheet.max_column
    # 获取性质列
    for i in range(3, 17):
        pro_cols.append(tar_sheet.cell(row=3, column=i).value)
    for i in range(4, row + 1):
        row_data = []
        for j in range(3, 17):
            row_data.append(tar_sheet.cell(row=i, column=j).value)
    
        pro_data.append(row_data)

    # 获取操作变量列
    for i in range(17, col + 1):
        op_cols.append(tar_sheet.cell(row=2, column=i).value)
    for i in range(4, row + 1):
        row_data = []
        for j in range(17, col + 1):
            row_data.append(tar_sheet.cell(row=i, column=j).value)
        
        op_data.append(row_data)
    
    return pro_data, pro_cols, op_data, op_cols


# 获取操作变量的中英文dict
def get_op_dict():
    eng_col = []
    chi_col = []

    wb = oxl.load_workbook(del_path)
    tar_sheet = wb.get_sheet_by_name("Sheet1")
    col = tar_sheet.max_column
    # 获取性质列
    for i in range(16, col):
        eng_col.append(tar_sheet.cell(row=2, column=i).value)
    for i in range(16, col):
        chi_col.append(tar_sheet.cell(row=3, column=i).value)

    return dict(zip(eng_col, chi_col))


# 获取操作变量取值范围
def get_range(path=tar_path):
    # 获取样本数据
    _, _, data, cols = load_tar_xml(path)
    data = np.array(data).astype(np.float)
    # 记录变量取值范围
    v_range = dict()
    # 取样本中数据的最值
    for i in range(len(cols)):
        v_ceil = max(data[:, i])
        v_floor = min(data[:, i])
        v_range.update({cols[i]: [v_floor, v_ceil]})

    return v_range


# 获取原始数据
def load_ori_xml(path):
    # 操作变量列
    op_cols = []
    op_data285 = []
    op_data313 = []
    # # 性质列
    # pro_cols = []
    # pro_data285 = []
    # pro_data313 = []
    # 样本行
    row285 = range(4, 44)
    row313 = range(45, 85)

    wb = oxl.load_workbook(path)
    # names = wb.get_sheet_names()
    # 操作变量sheet
    op_sheet = wb.get_sheet_by_name("操作变量")
    # # 原料sheet
    # pro_sheet1 = wb.get_sheet_by_name("原料")
    # # 产品sheet
    # pro_sheet2 = wb.get_sheet_by_name("产品")
    # # 待生吸附剂sheet
    # pro_sheet3 = wb.get_sheet_by_name("待生吸附剂")
    # # 再生吸附剂sheet
    # pro_sheet4 = wb.get_sheet_by_name("再生吸附剂")

    # 处理操作变量sheet
    col = op_sheet.max_column
    # 读取操作变量列(英文,中文就不处理了)
    for i in range(1, col + 1):
        op_cols.append(op_sheet.cell(row=2, column=i).value)
    # 读取数据
    for i in row285:
        temp_row = []
        for j in range(1, col + 1):
            temp_row.append(op_sheet.cell(row=i, column=j).value)
        op_data285.append(temp_row)
    for i in row313:
        temp_row = []
        for j in range(1, col + 1):
            temp_row.append(op_sheet.cell(row=i, column=j).value)
        op_data313.append(temp_row)

    # # 处理性质sheet
    # t1 = []
    # t2 = []
    # # 原料sheet
    # for i in row285:
    #     pro_cols.append(pro_sheet1.cell(row=1, column=i).value)
    #     t1.append(pro_sheet1.cell(row=2, column=i).value)
    #     t2.append(pro_sheet1.cell(row=3, column=i).value)
    # # 产品sheet
    # for i in row285:
    #     pro_cols.append(pro_sheet2.cell(row=2, column=i).value)
    #     t1.append(pro_sheet2.cell(row=3, column=i).value)
    #     t2.append(pro_sheet2.cell(row=4, column=i).value)
    # # 待生吸附剂sheet
    # for i in row285:
    #     pro_cols.append(pro_sheet3.cell(row=2, column=i).value)
    #     t1.append(pro_sheet3.cell(row=3, column=i).value)
    #     t2.append(pro_sheet3.cell(row=4, column=i).value)
    # # 再生吸附剂sheet
    # for i in row285:
    #     pro_cols.append(pro_sheet4.cell(row=2, column=i).value)
    #     t1.append(pro_sheet4.cell(row=3, column=i).value)
    #     t2.append(pro_sheet4.cell(row=4, column=i).value)
    # pro_data285.append(t1)
    # pro_data285.append(t2)

    return op_cols, op_data285, op_data313


# 预处理285,313操作变量数据
def preProcess(data, cols):
    resData = []
    # 获取操作变量范围
    v_range = get_range()
    # 数据整定
    data = np.array(data)
    # 跳过时间列
    for i in range(1, len(data[0])):
        cur_col_data = data[:, i].astype(np.float)
        # 1.删除全部为空的位点
        if len(np.nonzero(cur_col_data)[0]) > 0:
            # 2.处理部分数据为空值的位点，如果此列存在越0值则跳过
            if any(cur_col_data < 0) and any(cur_col_data > 0):
                pass
            else:
                col_av = np.average(cur_col_data)
                for j in range(len(cur_col_data)):
                    if j == 0:
                        cur_col_data[j] = col_av
            # 3.根据样本中变量范围，过滤原始数据
            p_data = []
            for j in cur_col_data:
                if j >= v_range[cols[i]][0] and j <= v_range[cols[i]][1]:
                    p_data.append(j)
            resData.append({cols[i]: np.average(np.array(p_data))})

            # # 4.根据拉益达准则去除异常值
            # p_res = layida(data[:, i].astype(np.float), cols[i])
            # resData.append(p_res)
        else:
            continue
    
    return resData


# 处理性质数据
def processProperty(data, cols):
    resData = []
    
    for v, k in zip(data, cols):
        resData.append(layida(v, k))

    return resData

# 拉一大准则
def layida(data, col):
    t_res = []
    p_av_mat = np.full_like(data, np.average(data))
    p_av = np.average(data)
    theta = math.sqrt(np.linalg.norm((p_av_mat - data), 2) / (len(data) - 1))
    for k in range(len(data)):
        if abs(data[k] - p_av) <= 3 * theta:
            t_res.append(data[k])
    return {col: np.average(np.array(t_res))}
    

#
def processOridata(data, cols):
    resData = []
    resCol = []
    # 数据整定
    data = np.array(data)
    # 跳过时间列
    for i in range(len(data[0])):
        cur_col_data = data[:, i].astype(np.float)
        # 1.删除全部为空的位点
        if len(np.nonzero(cur_col_data)[0]) > 0:
            # 2.处理部分数据为空值的位点，如果此列存在越0值则跳过
            if any(cur_col_data < 0) and any(cur_col_data > 0):
                pass
            else:
                col_av = np.average(cur_col_data)
                for j in range(len(cur_col_data)):
                    if j == 0:
                        cur_col_data[j] = col_av
            # 整理输出数据
            resCol.append(cols[i])
            resData.append(cur_col_data)
        else:
            print(cols[i])
            continue
    
    return resCol, resData


# 写出数据,xlwt库不支持多余256列的写入，因此使用csv格式导出
def output_data(data, path):
    keys = []
    values = []
    for d in data:
        keys.append(list(d.keys())[0])
        values.append(list(d.values())[0])
    
    # 输出csv
    pd.DataFrame(dict(zip(keys, values)), index=[0]).to_csv(path, index=False)
    


# %%
if __name__ == "__main__":
    # cols, data285, data313 = load_ori_xml(ori_path)
    # res285 = preProcess(data285, cols)
    # res313 = preProcess(data313, cols)
    # output_data(res285, r'C:\Users\wuziyang\Desktop\285(1).csv')
    # output_data(res313, r'C:\Users\wuziyang\Desktop\313(1).csv')

    pass
